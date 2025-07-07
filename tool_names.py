import os
import pandas as pd
from openpyxl import load_workbook

# === CONFIGURATION ===
PARENT_DIRECTORY = r"C:\Users\navak\Desktop\Hex 2025\Ticket Analyzer\Usecases"
OUTPUT_PY = 'a2_values.py'

# === DATA STRUCTURE ===
tool_names = []

# === PROCESS ALL FILES ===
total_files = 0
processed_files = 0
failed_files = 0

for root, dirs, files in os.walk(PARENT_DIRECTORY):
    for file in files:
        total_files += 1
        file_path = os.path.join(root, file)

        if file.endswith(('.xlsx', '.xlsm')):
            try:
                wb = load_workbook(filename=file_path, read_only=True)
                ws = wb.active
                value = ws['A2'].value
                if value:
                    tool_names.append(str(value).strip())
                wb.close()
                processed_files += 1
            except Exception as e:
                print(f"[ERROR] Could not open Excel file: {file_path} ({e})")
                failed_files += 1

        elif file.endswith(('.csv', '.csv.csv')):
            try:
                df = pd.read_csv(file_path, encoding='utf-8', engine='python')
                if not df.empty and df.shape[0] >= 1 and df.shape[1] >= 2:
                    value = df.iloc[0, 1]  # A2 in CSV means: first row, second column
                    if pd.notna(value):
                        tool_names.append(str(value).strip())
                processed_files += 1
            except Exception as e:
                print(f"[ERROR] Could not open CSV file: {file_path} ({e})")
                failed_files += 1

# === WRITE PYTHON MODULE ===
with open(OUTPUT_PY, 'w', encoding='utf-8') as f:
    f.write("tool_names = [\n")
    for tool in tool_names:
        safe_tool = str(tool).replace('"', '\\"')
        f.write(f'    "{safe_tool}",\n')
    f.write("]\n")

# === SUMMARY ===
print(f"=== SUMMARY ===")
print(f"Total files found: {total_files}")
print(f"Successfully processed: {processed_files}")
print(f"Failed to open: {failed_files}")
print(f"Data saved to {OUTPUT_PY}")
